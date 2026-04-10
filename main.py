"""
FastAPI backend for Amazon.de packaging dimension lookup.
Run with: uvicorn main:app --reload
"""
import io
import csv
import traceback
import asyncio
from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from scraper import run_search

app = FastAPI()


class SearchRequest(BaseModel):
    queries: list[str]        # one or more product queries
    max_variants: int = 3     # how many model variants to capture per query


@app.post("/api/search")
async def search(req: SearchRequest):
    all_rows = []
    for query in req.queries:
        try:
            rows = await run_search(query.strip(), max_variants=req.max_variants)
            for r in rows:
                r["query"] = query
            all_rows.extend(rows)
        except Exception as e:
            traceback.print_exc()
            all_rows.append({
                "query": query,
                "error": f"{type(e).__name__}: {e}" or "unknown error",
                "asin": "",
                "title": "",
                "model_codes": [],
                "pkg_dims": "",
                "pkg_weight": "",
                "pkg_volume": "",
            })
    return JSONResponse(content=all_rows)


@app.get("/api/export-csv")
async def export_csv(queries: str, max_variants: int = 3):
    """Export results as CSV. `queries` is newline-separated."""
    query_list = [q.strip() for q in queries.split("\n") if q.strip()]
    all_rows = []
    for query in query_list:
        try:
            rows = await run_search(query, max_variants=max_variants)
            for r in rows:
                r["query"] = query
            all_rows.extend(rows)
        except Exception as e:
            all_rows.append({"query": query, "error": str(e)})

    output = io.StringIO()
    fieldnames = ["query", "model_codes", "title", "asin", "url",
                  "pkg_dims", "pkg_weight", "pkg_volume", "error"]
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in all_rows:
        if "model_codes" in row:
            row["model_codes"] = ", ".join(row["model_codes"])
        writer.writerow(row)

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=packaging_data.csv"},
    )


@app.post("/api/export-excel")
async def export_excel(req: Request):
    """Build a styled Excel file from already-fetched results (sent as JSON body)."""
    rows = await req.json()

    wb = Workbook()
    ws = wb.active
    ws.title = "Packaging Data"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0066C0", end_color="0066C0", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    alt_fill = PatternFill(start_color="F5F8FF", end_color="F5F8FF", fill_type="solid")

    headers = ["Query", "Model Code(s)", "Product Title", "ASIN", "URL",
               "Pkg Dimensions", "Pkg Weight", "Pkg Volume"]
    col_widths = [25, 16, 50, 14, 40, 24, 16, 14]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = w

    ws.auto_filter.ref = f"A1:H1"
    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(rows, 2):
        model_codes = row.get("model_codes", [])
        if isinstance(model_codes, list):
            model_codes = ", ".join(model_codes)
        values = [
            row.get("query", ""),
            model_codes,
            row.get("title", ""),
            row.get("asin", ""),
            row.get("url", ""),
            row.get("pkg_dims", ""),
            row.get("pkg_weight", ""),
            row.get("pkg_volume", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 3))
            if row_idx % 2 == 0:
                cell.fill = alt_fill
        # Make URL a hyperlink
        url_cell = ws.cell(row=row_idx, column=5)
        if url_cell.value:
            url_cell.hyperlink = url_cell.value
            url_cell.font = Font(color="0066C0", underline="single")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=packaging_data.xlsx"},
    )


@app.get("/", response_class=HTMLResponse)
async def ui():
    with open("index.html", encoding="utf-8") as f:
        return f.read()
